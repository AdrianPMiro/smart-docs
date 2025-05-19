# generators/excel_generator.py

import openpyxl
from openpyxl.utils.cell import get_column_letter
from pathlib import Path
from typing import Dict, Optional, Tuple
import unicodedata
import re


# --- Función de Normalización (MENOS AGRESIVA) ---
def _excel_norm_label(text: str) -> str:
    """
    Normaliza etiquetas para búsqueda: minúsculas, sin acentos, espacios simples.
    ¡IMPORTANTE! NO quita puntuación por defecto (ej. ':').
    """
    if not isinstance(text, str):
        text = str(text)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()  # Solo minúsculas
    # text = re.sub(r"[:]", "", text) # Descomenta si QUIERES quitar los dos puntos
    return re.sub(r"\s+", " ", text).strip()  # Normalizar espacios


# --- Función de Búsqueda de Etiquetas (con COINCIDENCIA EXACTA y LOGS) ---
def _find_label_cell_and_get_value_target(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        label_pattern_norm_to_find: str,  # La etiqueta normalizada EXACTA que buscamos
        search_cols_limit: int = 10,
        value_col_offset: int = 1
) -> Optional[Tuple[str, str]]:
    """
    Busca una celda cuyo contenido normalizado COINCIDA EXACTAMENTE con label_pattern_norm_to_find.
    Devuelve la coordenada de la celda valor y el texto original de la etiqueta.
    """
    # print(f"DEBUG: Buscando etiqueta normalizada exacta: '{label_pattern_norm_to_find}'") # Log de qué buscamos
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        effective_cols_limit = min(search_cols_limit, sheet.max_column)
        for cell in row[:effective_cols_limit]:
            if cell.value is not None:
                cell_original_text = str(cell.value)
                cell_text_norm = _excel_norm_label(cell_original_text)  # Usar la misma normalización

                # --- LOG DETALLADO DE COMPARACIÓN (Descomentar si es necesario) ---
                # if row_idx < 20 and cell.column < 5: # Limitar logs a las primeras filas/columnas
                #     print(f"  DEBUG: Celda {cell.coordinate}='{cell_original_text}' -> Norm='{cell_text_norm}' | Comparando con '{label_pattern_norm_to_find}'")
                # --------------------------------------------------------------------

                # Comparación EXACTA
                if label_pattern_norm_to_find == cell_text_norm:
                    print(
                        f"  DEBUG: ¡COINCIDENCIA EXACTA! Patrón '{label_pattern_norm_to_find}' encontrado en celda {cell.coordinate}")  # Log de éxito
                    target_col_idx = cell.column + value_col_offset
                    target_row_idx = cell.row
                    if target_col_idx <= sheet.max_column:
                        target_cell_coordinate = f"{get_column_letter(target_col_idx)}{target_row_idx}"
                        return target_cell_coordinate, cell_original_text.strip()
                    else:
                        print(
                            f"WARN: Etiqueta encontrada en {cell.coordinate}, pero columna valor {target_col_idx} excede límite.")
                        return None  # No podemos obtener coordenada válida

    # print(f"DEBUG: Etiqueta normalizada '{label_pattern_norm_to_find}' NO ENCONTRADA.") # Log si no se encuentra
    return None


# --- Función Principal del Generador ---
def excel_generator(
        datos_nuevos: Dict[str, str],
        plantilla_excel_path: Path,
        salida_excel_path: Path,
        sheet_to_modify_idx: int = 0,
        value_column_offset: int = 1
) -> None:
    print(f"ℹ️  [excel_generator v3] Iniciando generación Excel desde plantilla: {plantilla_excel_path}")

    # --- Mapa: Etiqueta NORMALIZADA a Buscar -> Clave en datos_nuevos ---
    # !!! === AJUSTA ESTAS CLAVES === !!!
    # Aplica _excel_norm_label() al texto EXACTO de tus etiquetas en Excel.
    # Incluye los dos puntos ':' si están presentes en el Excel.
    LABEL_SEARCH_TO_DATA_KEY = {
        _excel_norm_label("NIF:"): "nif",  # Probablemente tiene ':' en Excel
        _excel_norm_label("Nombre/Razón Social:"): "nombre",  # Probablemente tiene ':' y '/'
        _excel_norm_label("Primer Apellido:"): "apellido1",  # Probablemente tiene ':'
        _excel_norm_label("Segundo Apellido:"): "apellido2",  # Probablemente tiene ':'
        _excel_norm_label("Tipo vía:"): "tipo_via",  # Probablemente tiene ':'
        _excel_norm_label("Nombre vía:"): "nom_via",  # Probablemente tiene ':'
        _excel_norm_label("Móvil:"): "telefono",  # Probablemente tiene ':'
        _excel_norm_label("Telefono:"): "telefono",  # Alternativa
        _excel_norm_label("CP:"): "cp",  # Probablemente tiene ':'
        _excel_norm_label("Código Postal:"): "cp",  # Alternativa
        _excel_norm_label("Localidad:"): "localidad",  # Probablemente tiene ':'
        _excel_norm_label("Población:"): "localidad",  # Alternativa
        _excel_norm_label("Provincia:"): "provincia",  # Probablemente tiene ':'
        # Añade la etiqueta para la dirección COMPLETA si la necesitas reemplazar como un todo
        # _excel_norm_label("Dirección:"): "direccion", # Asumiendo que 'direccion' existe en datos_nuevos
        # ... añade más mapeos según tu plantilla exacta ...
    }
    print("--- Mapa de Búsqueda (Etiqueta Normalizada -> Clave Datos) ---")
    for k, v in LABEL_SEARCH_TO_DATA_KEY.items(): print(f"'{k}' -> '{v}'")
    print("-------------------------------------------------------------")

    try:
        workbook = openpyxl.load_workbook(plantilla_excel_path)
        # ... (resto de la lógica para seleccionar la hoja, manejar .xls como antes) ...
        if sheet_to_modify_idx < len(workbook.sheetnames):
            sheet = workbook.worksheets[sheet_to_modify_idx]
            print(f"ℹ️  [excel_generator v3] Modificando hoja: '{sheet.title}'")
        else:  # Fallback
            sheet = workbook.active
            if not sheet: raise ValueError("No se pudo seleccionar hoja activa.")
            print(f"⚠️  [excel_generator v3] Índice de hoja inválido, usando hoja activa: '{sheet.title}'")

    except Exception as e:
        # ... (manejo de errores al abrir, incluyendo el de .xls) ...
        if "does not support the old .xls file format" in str(e):
            print(f"❌ [excel_generator v3] Plantilla .xls no soportada. Convierte a .xlsx.")
        else:
            print(f"❌ [excel_generator v3] Error al abrir/leer plantilla: {e}")
        return

    modified_cells_count = 0
    fields_written = set()

    # Iterar sobre el MAPA de etiquetas que queremos encontrar y reemplazar
    for label_to_find_norm, data_key in LABEL_SEARCH_TO_DATA_KEY.items():
        print(
            f"--- Procesando campo: '{data_key}' (buscando etiqueta norm: '{label_to_find_norm}') ---")  # Log principal

        target_info = _find_label_cell_and_get_value_target(
            sheet, label_to_find_norm, value_col_offset=value_column_offset
        )

        if target_info:
            target_cell_coord, found_label_text = target_info
            print(f"  Etiqueta encontrada ('{found_label_text}'), celda valor objetivo: {target_cell_coord}")

            if data_key in datos_nuevos:
                nuevo_valor = datos_nuevos[data_key]

                # Limpieza del valor nuevo (ej. quitar .0)
                if isinstance(nuevo_valor, str) and nuevo_valor.endswith(".0"):
                    try:
                        if float(nuevo_valor) == int(float(nuevo_valor)): nuevo_valor = nuevo_valor[:-2]
                    except ValueError:
                        pass

                    # Escribir el nuevo valor
                try:
                    current_value_in_target = sheet[target_cell_coord].value
                    # Convertir ambos a string para comparación segura
                    if str(current_value_in_target or "") != str(nuevo_valor or ""):
                        sheet[target_cell_coord] = nuevo_valor  # Escribir valor (puede ser número o string)
                        print(f"    -> VALOR ESCRITO: '{nuevo_valor}' en celda {target_cell_coord}")
                        modified_cells_count += 1
                        fields_written.add(data_key)
                    else:
                        print(f"    -> Valor en {target_cell_coord} ya es '{nuevo_valor}'. No se modifica.")

                except Exception as e:
                    print(f"⚠️  [excel_generator v3] Error al escribir en celda {target_cell_coord}. Error: {e}")
            else:
                print(
                    f"  -> ADVERTENCIA: Clave de datos '{data_key}' no encontrada en 'datos_nuevos'. No se puede escribir valor.")
        else:
            print(f"  -> ETIQUETA NO ENCONTRADA en la hoja para '{data_key}'.")

    print(f"\n--- Resumen Final de Escritura ---")
    # ... (impresión de resumen como antes) ...
    print(f" Celdas modificadas: {modified_cells_count}")
    print(f" Campos de 'datos_nuevos' escritos: {fields_written if fields_written else 'Ninguno'}")
    missing_keys = set(LABEL_SEARCH_TO_DATA_KEY.values()) - fields_written
    if missing_keys: print(f" Campos mapeados pero NO escritos: {missing_keys}")
    print(f"-------------------------------\n")

    try:
        workbook.save(salida_excel_path)
        print(f"✅ [excel_generator v3] Documento Excel modificado guardado en: {salida_excel_path}")
    except Exception as e:
        print(
            f"❌ [excel_generator v3] Error al guardar el documento Excel modificado: {salida_excel_path}. Detalle: {e}")